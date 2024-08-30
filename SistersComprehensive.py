"""This function writes sisters data only in the spatial
domain of interest.
Input are the analysis folder and xlsx comprehensive file.
"""


import xlsxwriter
from openpyxl import load_workbook
import datetime


class SistersComprehensive:
    """Only class, does all the job"""
    def __init__(self, analysis_folder):

        idx2start  =  None
        idx2stop   =  None
        nucs_list  =  []
        wb         =  load_workbook(analysis_folder + '/ComprehensiveBurstingData.xlsx').active

        for brws_idx in range(1, 10000):
            try:
                if wb.cell(row=brws_idx, column=1).value[:4] != "Nuc_":
                    brws_idx  +=  1
                else:
                    idx2start  =  brws_idx
                    break
            except TypeError:
                brws_idx += 1

        for brws_idx2 in range(idx2start, 10000):
            try:
                if wb.cell(row=brws_idx2, column=1).value[:4] == "Nuc_":
                    brws_idx2  +=  1
                else:
                    idx2stop  =  brws_idx2 - 1
                    break
            except TypeError:
                idx2stop  =  brws_idx2 - 1
                break

        for kk in range(idx2start, idx2stop + 1):
            nucs_list.append(wb.cell(row=kk, column=1).value)

        wb2cp       =  load_workbook(analysis_folder + '/AlleleIntensity.xlsx')
        colmn_refs  =  [1]
        for nuc_tag in nucs_list:
            for x in range(2, 10000, 3):
                if wb2cp[wb2cp.sheetnames[0]].cell(row=1, column=x).value == nuc_tag:
                    colmn_refs  +=  x, x + 1, x + 2
                    break

        lst_row  =  1
        while wb2cp[wb2cp.sheetnames[0]].cell(row=lst_row, column=1).value is not None:
            lst_row  +=  1

        book    =  xlsxwriter.Workbook(analysis_folder + "/AlleleIntensitySpatialSelected.xlsx")                                                                  # write results
        sheet1  =  book.add_worksheet("Ints")
        sheet2  =  book.add_worksheet("Bckg")
        sheet3  =  book.add_worksheet("Ints by Bckg")

        for cc, colmn_ref in enumerate(colmn_refs):
            for rr in range(1, lst_row):
                sheet1.write(rr - 1, cc, wb2cp[wb2cp.sheetnames[0]].cell(row=rr, column=colmn_ref).value)
                sheet2.write(rr - 1, cc, wb2cp[wb2cp.sheetnames[1]].cell(row=rr, column=colmn_ref).value)
                sheet3.write(rr - 1, cc, wb2cp[wb2cp.sheetnames[2]].cell(row=rr, column=colmn_ref).value)

        sheet1.write(lst_row + 2, 0, "Date")
        sheet1.write(lst_row + 2, 1, datetime.datetime.now().strftime("%d-%b-%Y"))

        book.close()
