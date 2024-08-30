"""This function generates gallery of traces intensity
2 spots per nucleus data. In each plot the two intensity
traces are be shown in different colors of course.
"""

import numpy as np
import time
from openpyxl import load_workbook
# import pyqtgraph as pg


class TwoSpotsGallery:
    def __init__(self, analysis_folder):

        wb  =  load_workbook(analysis_folder + '/AlleleIntensity.xlsx')
        wb  =  wb[wb.sheetnames[0]]

        sists_trcks  =  np.zeros((2, wb.max_row - 4, int((wb.max_column - 1) / 3)), dtype=int)
        tags_list    =  list()

        for tt in range(sists_trcks.shape[2]):
            tags_list.append(wb.cell(row=1, column=2 + 3 * tt).value)
            for k in range(sists_trcks.shape[1]):
                sists_trcks[0, k, tt]  =  wb.cell(row=k + 2, column=2 + 3 * tt).value
                sists_trcks[1, k, tt]  =  wb.cell(row=k + 2, column=3 + 3 * tt).value

        y_sup    =  sists_trcks.max()
        n_rows   =  6
        n_cols   =  7
        num_win  =  len(tags_list) // (n_cols * n_rows) + 1
        for win_idxs in range(num_win):
            time.sleep(1)
            # str_win1  =  "win" + str(win_idxs) + "  =  pg.GraphicsWindow()"
            str_win1  =  "win" + str(win_idxs) + "  =  pg.GraphicsLayoutWidget()"
            str_win2  =  "win" + str(win_idxs) + ".setWindowTitle('Transcriptional Traces " + str(win_idxs + 1) + "')"
            str_win3  =  "win" + str(win_idxs) + ".showMaximized()"
            exec(str_win1)
            exec(str_win2)
            exec(str_win3)

            idx_name = 0
            for idx_r in range(n_rows):
                for idx_c in range(n_cols):
                    str_addplot  =  "p" + str(idx_name) +  "= win" + str(win_idxs) + ".addPlot(row=" + str(idx_r) + ", col=" + str(idx_c) + ")"
                    exec(str_addplot)
                    idx_name  +=  1

            for k in range(n_cols * n_rows):
                if k + win_idxs * n_cols * n_rows <= len(tags_list) - 1:
                    str_cmnd1  =  "p" + str(k) + ".plot(sists_trcks[0, :, k  + win_idxs * n_cols * n_rows], pen='r', symbol='o', symbolSize=2)"
                    str_cmnd1_2  =  "p" + str(k) + ".plot(sists_trcks[1, :, k  + win_idxs * n_cols * n_rows], pen='g', symbol='o', symbolSize=2)"
                    str_cmnd2  =  "p" + str(k) + ".setYRange(0, y_sup)"
                    str_cmnd3  =  "tag_text" + str(k) + " = pg.TextItem('tag = ' + tags_list[k + win_idxs * n_cols * n_rows], color='g')"
                    str_cmnd4  =  "tag_text" + str(k) + ".setPos(1, y_sup)"
                    str_cmnd5  =  "p" + str(k) + ".addItem(tag_text" + str(k) + ")"
                    exec(str_cmnd1)
                    exec(str_cmnd1_2)
                    exec(str_cmnd2)
                    exec(str_cmnd3)
                    exec(str_cmnd4)
                    exec(str_cmnd5)
                else:
                    break
        print(StrangePatch)
