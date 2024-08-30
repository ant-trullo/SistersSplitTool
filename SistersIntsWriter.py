"""This function writes results of the sister tracking.
"""

import datetime
import numpy as np
from skimage.morphology import label
from skimage.measure import regionprops_table
from skimage.segmentation import expand_labels
import xlsxwriter
from openpyxl import load_workbook
# from openpyxl import load_workbook
# from PyQt5 import QtWidgets

import ServiceWidgets


def sists_3d(sists_3D_coords, frame):
    """Build 3d all the present spots in a single frame"""
    sists_3d        =  np.zeros(sists_3D_coords[-1, (2, 3, 4)], dtype=np.uint16)
    sub_3D_coords  =  sists_3D_coords[sists_3D_coords[:, 1] == frame]
    sists_3d[sub_3D_coords[:, 2], sub_3D_coords[:, 3], sub_3D_coords[:, 4]]  =  sub_3D_coords[:, 0]
    return sists_3d


class SistersIntsWriter:
    """Write xlsx file and the npy matrix of the results"""
    def __init__(self, analysis_folder, sists_ints, sists_bckg, sists_3D_coords, spts_idxs, tags2rm, dist_thr, software_version):

        old_book              =  load_workbook(analysis_folder + '/journal.xlsx')
        old_book_s            =  old_book[old_book.sheetnames[0]]
        pix_size, pix_size_Z  =  old_book_s.cell(row=16, column=2).value, old_book_s.cell(row=17, column=2).value

        steps     =  sists_ints.shape[2]
        spts_trk  =  np.load(analysis_folder + '/spots_tracked.npy')              # load nuclei-tracked spots
        book      =  xlsxwriter.Workbook(analysis_folder + "/AlleleIntensity.xlsx")                                                                  # write results
        sheet1    =  book.add_worksheet("Ints")
        sheet2    =  book.add_worksheet("Bckg")
        sheet3    =  book.add_worksheet("Ints by Bckg")

        for uu in tags2rm:
            spts_idxs  =  np.delete(spts_idxs, np.where(spts_idxs == uu))

        sheet1.write(0, 0, "Frame")
        sheet2.write(0, 0, "Frame")
        sheet3.write(0, 0, "Frame")
        for tt in range(steps):
            sheet1.write(tt + 1, 0, tt)
            sheet2.write(tt + 1, 0, tt)
            sheet3.write(tt + 1, 0, tt)

        for cc in range(steps):
            for cnt, k in enumerate(spts_idxs):
                sists3d  =  sists_3d(sists_3D_coords, cc)
                sheet1.write(0, 3 * cnt + 1, "Nuc_" + str(k))
                sheet2.write(0, 3 * cnt + 1, "Nuc_" + str(k))
                sheet3.write(0, 3 * cnt + 1, "Nuc_" + str(k))
                ref_crd  =  np.where(spts_idxs == k)[0][0]
                sheet1.write(cc + 1, 3 * cnt + 1, sists_ints[ref_crd, 0, cc])
                sheet2.write(cc + 1, 3 * cnt + 1, sists_bckg[ref_crd, 0, cc])
                if sists_bckg[ref_crd, 0, cc] != 0:
                    sheet3.write(cc + 1, 3 * cnt + 1, sists_ints[ref_crd, 0, cc] / sists_bckg[ref_crd, 0, cc])
                else:
                    sheet3.write(cc + 1, 3 * cnt + 1, 0)
                sheet1.write(cc + 1, 3 * cnt + 2, sists_ints[ref_crd, 1, cc])
                sheet2.write(cc + 1, 3 * cnt + 2, sists_bckg[ref_crd, 1, cc])
                if sists_bckg[ref_crd, 1, cc] != 0:
                    sheet3.write(cc + 1, 3 * cnt + 2, sists_ints[ref_crd, 1, cc] / sists_bckg[ref_crd, 1, cc])
                else:
                    sheet3.write(cc + 1, 3 * cnt + 2, 0)

                sists_tags  =  sists3d  *  (spts_trk[cc] == k)
                sists_tags  =  np.unique(sists_tags[sists_tags != 0])
                if sists_tags.size == 2:
                    sub_3D_coords   =  sists_3D_coords[sists_3D_coords[:, 1] == cc]
                    sub_3D_coords0  =  sub_3D_coords[sub_3D_coords[:, 0] == sists_tags[0]]
                    sub_3D_coords1  =  sub_3D_coords[sub_3D_coords[:, 0] == sists_tags[1]]
                    ctrs0           =  [sub_3D_coords0[:, 2].mean(), sub_3D_coords0[:, 3].mean(), sub_3D_coords0[:, 4].mean()]
                    ctrs1           =  [sub_3D_coords1[:, 2].mean(), sub_3D_coords1[:, 3].mean(), sub_3D_coords1[:, 4].mean()]
                    dist            =  np.sqrt(((ctrs0[0] - ctrs1[0]) * pix_size_Z) ** 2 + ((ctrs0[1] - ctrs1[1]) * pix_size) ** 2 + ((ctrs0[2] - ctrs1[2]) * pix_size) ** 2)
                    sheet1.write(cc + 1, 3 * cnt + 3, dist)
                    sheet3.write(cc + 1, 3 * cnt + 3, dist)
                    sists_ints[ref_crd, 2, cc]  =  dist

        sheet1.write(steps + 2, 0, "Analysis Folder")
        sheet1.write(steps + 3, 0, analysis_folder)

        sheet1.write(steps + 2, 1, "dist-thr")
        sheet1.write(steps + 3, 1, dist_thr)

        sheet1.write(steps + 2, 3, "Date")
        sheet1.write(steps + 3, 3, datetime.datetime.now().strftime("%d-%b-%Y"))

        sheet1.write(steps + 2, 4, "Software-version")
        sheet1.write(steps + 3, 4, software_version)

        book.close()

        np.save(analysis_folder + '/sists_ints.npy', sists_ints)
        np.save(analysis_folder + '/sists_3D_coords.npy', sists_3D_coords)
        np.save(analysis_folder + '/sists_bckg.npy', sists_bckg)


class LoadSistsTrack:
    """Load the saved matrices with all the values of the sisters tracking"""
    def __init__(self, analysis_folder):

        self.sists_ints       =  np.load(analysis_folder + '/sists_ints.npy')
        self.sists_3D_coords  =  np.load(analysis_folder + '/sists_3D_coords.npy')
        self.sists_trck       =  np.zeros((self.sists_3D_coords[-1, 1], self.sists_3D_coords[-1, 3], self.sists_3D_coords[-1, 4]), dtype=np.uint32)
        for jj in self.sists_3D_coords[:-1]:
            self.sists_trck[jj[1], jj[3], jj[4]]  =  jj[0]


class LoadSistsTrackDistThr:
    """Load the saved matrices with all the values of the sisters tracking"""
    def __init__(self, analysis_folder):

        wb    =  load_workbook(analysis_folder + '/AlleleIntensity.xlsx')
        wb_s  =  wb[wb.sheetnames[0]]
        for rr in range(1, wb_s.max_row + 1):
            if wb_s.cell(rr, 2).value == "dist-thr":
                self.dist_thr  =  wb_s.cell(rr + 1, 2).value
                break


class LabelLeft:
    """This function labels leftovers spots from sisters tracking"""
    def __init__(self, spts_trck, sists_trck):

        if len(spts_trck.shape) == 3:
            spts_left  =  np.zeros_like(spts_trck)
            for tt in range(spts_left.shape[0]):
                spts_left[tt]  =  label(spts_trck[tt] * (1 - np.sign(sists_trck[tt])), connectivity=1)

            self.spts_left  =  spts_left

        elif len(spts_trck.shape) == 2:
            self.spts_left   =  label(spts_trck * (1 - np.sign(sists_trck)), connectivity=1)


class WriteMomentaryInfo:
    """This function writes the distance info in xlsx file as before the background calculation"""
    def __init__(self, analysis_folder, dist_thr):

        book    =  xlsxwriter.Workbook(analysis_folder + "/AlleleIntensity.xlsx")                         # write results
        sheet1  =  book.add_worksheet("Ints")
        sheet1.write(2, 1, "dist-thr")
        sheet1.write(3, 1, dist_thr)
        book.close()


class ReadTags2Rm:
    """This function get from the analysis folder the list of the nuclei tags removed during the sister analysis"""
    def __init__(self, analysis_folder):

        tot_tags      =  np.load(analysis_folder + '/spots_tracked.npy')                        # total tags from the tracked spots
        tot_tags      =  list(np.unique(tot_tags))[1:]
        allele_book   =  load_workbook(analysis_folder + '/AlleleIntensity.xlsx')               # read AlleleIntensity excel file
        allele_sheet  =  allele_book[allele_book.sheetnames[0]]
        tags_used     =  list()
        colmns        =  np.arange(1, allele_sheet.max_column, 3) + 1
        [tags_used.append(int(allele_sheet.cell(row=1, column=x).value[4:])) for x in colmns]   # get the list of the tags used
        [tot_tags.remove(tag_used) for tag_used in tags_used]                                   # remove the used tags from the total tags

        self.tags_removed  =  tot_tags                                                          # list of tags removed during the analysi


class SistsIntsDist:
    """Only class, does all the job"""
    def __init__(self, analysis_folder, green4D, tags2rm, software_version, pix_size, pix_size_Z, dist_thr):

        # raw_data          =  AnalysisLoader.RawData(analysis_folder)
        spts_trck         =  np.load(analysis_folder + '/spots_tracked.npy')                    # load data from the analysis folder
        sists             =  LoadSistsTrack(analysis_folder)
        sists_3D_coords   =  np.load(analysis_folder + '/sists_3D_coords.npy')
        steps             =  spts_trck.shape[0]
        spts_tags         =  np.unique(spts_trck[spts_trck != 0])                               # spots tracked tags (corresponfing to nuclei tags)
        sists_ints        =  np.zeros((spts_tags.size, 3, spts_trck.shape[0]))                  # initialize final matrix
        zlen, xlen, ylen  =  sists_3D_coords[-1, 2:]                                            # 3D shape
        sists_3D_coords   =  sists_3D_coords[:-1, :]                                            # clean sists_3D_coords info (remove size info)
        tot_sists_tags    =  np.unique(sists_3D_coords[:, 0])                                   # list of all the allele tags
        sists_bkgd_ref    =  np.zeros((tot_sists_tags.size, steps), np.uint32)
        sists_bckg        =  np.zeros((spts_tags.size, 2, steps))

        pbar  =  ServiceWidgets.ProgressBarDouble(total1=steps, total2=spts_tags.size)
        pbar.show()
        pbar.update_progressbar1(0)
        pbar.update_progressbar2(0)

        for t in range(steps):
            # print(t)
            pbar.update_progressbar1(t)
            sists_3D_coords_sub  =  sists_3D_coords[sists_3D_coords[:, 1] == t]            # isolate from coordinate matrix all the spots coordinate of the time frame
            spt_3d               =  np.zeros((zlen, xlen, ylen), dtype=np.uint16)
            for mm in sists_3D_coords_sub:                                                 # recrete 3D spots from coordinates
                spt_3d[mm[2], mm[3], mm[4]]  =  mm[0]

            aze_3   =  expand_labels(spt_3d, distance=3)                                   # expande labels with iterations 3
            aze_5   =  expand_labels(aze_3, distance=3) * (1 - np.sign(aze_3))             # expand with iterations 2 over iterations 3 and remove iterations 3 (you keep the shell)
            rgp_5   =  regionprops_table(aze_5, green4D[t], properties=["label", "intensity_image", "area"])   # regionprops of the shell
            for uu, lb in enumerate(rgp_5["label"]):                                        # organize bckg info
                ref_idx                     =  np.where(tot_sists_tags == lb)[0][0]         # find the proper label column
                sists_bkgd_ref[ref_idx, t]  =  np.sum(rgp_5["intensity_image"][uu]) / rgp_5["area"][uu]     # introduce the background value for the tag and time

        for c, spts_tag in enumerate(spts_tags):
            # print(c)
            pbar.update_progressbar2(c)
            spt         =  (spts_trck == spts_tag)                                                                          # isolate tracked spot tag
            sists_tags  =  sists.sists_trck * spt
            sists_tags  =  np.unique(sists_tags[sists_tags != 0])                                                           # identify allele tags for the spots
            if sists_tags.size == 2:                                                                                        # in case you have both allele present all along the evolution
                for ss in range(steps):                                                                                     # for each time step
                    bff_coords            =  sists_3D_coords[sists_3D_coords[:, 1] == ss]                                   # isolate coordinates of all the spots of the frame
                    bff_coords1           =  bff_coords[bff_coords[:, 0] == sists_tags[0]]                                  # among these, isolate coordinates of only one sister
                    sists_ints[c, 0, ss]  =  np.sum(green4D[ss, bff_coords1[:, 2], bff_coords1[:, 3], bff_coords1[:, 4]])   # intensity of the spot taken from raw data
                    ref_idx_a             =  np.where(tot_sists_tags == sists_tags[0])[0][0]                                # find the corresponding index in the background matrix
                    sists_bckg[c, 0, ss]  =  sists_bkgd_ref[ref_idx_a, ss]                                                  # write in the final matrix in corrispondence of the matrix intensity (tags are alligned)
                    ref_idx_b             =  np.where(tot_sists_tags == sists_tags[1])[0][0]                                # same for the second allele
                    bff_coords2           =  bff_coords[bff_coords[:, 0] == sists_tags[1]]
                    sists_ints[c, 1, ss]  =  np.sum(green4D[ss, bff_coords2[:, 2], bff_coords2[:, 3], bff_coords2[:, 4]])
                    sists_bckg[c, 1, ss]  =  sists_bkgd_ref[ref_idx_b, ss]
                    if sists_ints[c, 0, ss] != 0 and sists_ints[c, 1, ss] != 0:                                                         # in frames with allele present
                        ctrs_a                =  [bff_coords1[:, 2].mean(), bff_coords1[:, 3].mean(), bff_coords1[:, 4].mean()]         # calculate the centroid of both
                        ctrs_b                =  [bff_coords2[:, 2].mean(), bff_coords2[:, 3].mean(), bff_coords2[:, 4].mean()]
                        dist                  =  np.sqrt(((ctrs_a[0] - ctrs_b[0]) * pix_size_Z) ** 2 + ((ctrs_a[1] - ctrs_b[1]) * pix_size) ** 2 + ((ctrs_a[2] - ctrs_b[2]) * pix_size) ** 2)   # take distance considering that z and xy steps are different
                        sists_ints[c, 2, ss]  =  dist                                                                           # write in the matrix

            elif sists_tags.size == 1:
                for ss in range(steps):
                    ref_idx_0             =  np.where(tot_sists_tags == sists_tags[0])[0][0]
                    bff_coords            =  sists_3D_coords[sists_3D_coords[:, 1] == ss]
                    bff_coords1           =  bff_coords[bff_coords[:, 0] == sists_tags[0]]
                    sists_ints[c, 0, ss]  =  np.sum(green4D[ss, bff_coords1[:, 2], bff_coords1[:, 3], bff_coords1[:, 4]])
                    sists_bckg[c, 0, ss]  =  sists_bkgd_ref[ref_idx_0, ss]

        pbar.close()

        book    =  xlsxwriter.Workbook(analysis_folder + "/AlleleIntensity.xlsx")                                                                  # write results
        sheet1  =  book.add_worksheet("Ints")
        sheet2    =  book.add_worksheet("Bckg")
        sheet3    =  book.add_worksheet("Ints by Bckg")

        sheet1.write(0, 0, "Frame")
        sheet2.write(0, 0, "Frame")
        sheet3.write(0, 0, "Frame")
        for tt in range(steps):
            sheet1.write(tt + 1, 0, tt)
            sheet2.write(tt + 1, 0, tt)
            sheet3.write(tt + 1, 0, tt)

        clm_idx  =  0
        for cnt, spts_tag in enumerate(spts_tags):
            if spts_tag not in tags2rm:
                for cc in range(steps):
                    sheet1.write(0, 3 * clm_idx + 1, "Nuc_" + str(spts_tag))
                    sheet1.write(cc + 1, 3 * clm_idx + 1, sists_ints[cnt, 0, cc])
                    sheet1.write(cc + 1, 3 * clm_idx + 2, sists_ints[cnt, 1, cc])
                    sheet1.write(cc + 1, 3 * clm_idx + 3, np.nan_to_num(sists_ints[cnt, 2, cc]))
                    sheet2.write(0, 3 * clm_idx + 1, "Nuc_" + str(spts_tag))
                    sheet2.write(cc + 1, 3 * clm_idx + 1, sists_bckg[cnt, 0, cc])
                    sheet2.write(cc + 1, 3 * clm_idx + 2, sists_bckg[cnt, 1, cc])
                    sheet3.write(0, 3 * clm_idx + 1, "Nuc_" + str(spts_tag))
                    sheet3.write(cc + 1, 3 * clm_idx + 1, np.nan_to_num(sists_ints[cnt, 0, cc] / sists_bckg[cnt, 0, cc]))
                    sheet3.write(cc + 1, 3 * clm_idx + 2, np.nan_to_num(sists_ints[cnt, 1, cc] / sists_bckg[cnt, 1, cc]))
                    sheet3.write(cc + 1, 3 * clm_idx + 3, np.nan_to_num(sists_ints[cnt, 2, cc]))

                clm_idx  +=  1

        sheet1.write(steps + 2, 0, "Analysis Folder")
        sheet1.write(steps + 3, 0, analysis_folder)

        sheet1.write(steps + 2, 1, "dist-thr")
        sheet1.write(steps + 3, 1, dist_thr)

        sheet1.write(steps + 2, 3, "Date")
        sheet1.write(steps + 3, 3, datetime.datetime.now().strftime("%d-%b-%Y"))

        sheet1.write(steps + 2, 4, "Software-version")
        sheet1.write(steps + 3, 4, software_version)

        book.close()

        self.sists_ints  =  sists_ints
        self.sists_bckg  =  sists_bckg


# class ProgressBarDouble(QtWidgets.QWidget):
#     def __init__(self, parent=None, total1=20, total2=20):
#         super().__init__(parent)
#         self.name_line1  =  QtWidgets.QLineEdit()
#
#         self.progressbar1  =  QtWidgets.QProgressBar()
#         self.progressbar1.setMinimum(0)
#         self.progressbar1.setMaximum(total1)
#
#         self.progressbar2  =  QtWidgets.QProgressBar()
#         self.progressbar2.setMinimum(0)
#         self.progressbar2.setMaximum(total2)
#
#         main_layout  =  QtWidgets.QGridLayout()
#         main_layout.addWidget(self.progressbar1, 0, 0)
#         main_layout.addWidget(self.progressbar2, 1, 0)
#
#         self.setLayout(main_layout)
#         self.setWindowTitle("Progress")
#         self.setGeometry(500, 300, 300, 50)
#
#     def update_progressbar1(self, val1):
#         self.progressbar1.setValue(val1)
#         QtWidgets.qApp.processEvents()
#
#     def update_progressbar2(self, val2):
#         self.progressbar2.setValue(val2)
#         QtWidgets.qApp.processEvents()
